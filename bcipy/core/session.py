# mypy: disable-error-code="arg-type"
"""Helper functions for managing and parsing session.json data."""
import csv
import itertools
import json
import logging
import os
import sqlite3
from dataclasses import dataclass, fields
from typing import Any, Dict, List, Optional, Union, Iterator

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import BORDER_THIN, Border, Side
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.worksheet.worksheet import Worksheet

from bcipy.config import (DEFAULT_ENCODING, DEFAULT_PARAMETERS_FILENAME,
                          SESSION_DATA_FILENAME, SESSION_SUMMARY_FILENAME, SESSION_LOG_FILENAME)
from bcipy.io.load import load_json_parameters
from bcipy.task.data import Session

# Configure logging
logger = logging.getLogger(SESSION_LOG_FILENAME)

BLACK = COLOR_INDEX[0]
WHITE = COLOR_INDEX[1]
YELLOW = COLOR_INDEX[5]


def read_session(file_name: str = SESSION_DATA_FILENAME) -> Session:
    """Read session data from a JSON file.

    Args:
        file_name (str, optional): Path to the session data file. 
            Defaults to SESSION_DATA_FILENAME.

    Returns:
        Session: A Session object containing the parsed data.

    Raises:
        FileNotFoundError: If the specified file does not exist.
        json.JSONDecodeError: If the file contains invalid JSON.
    """
    with open(file_name, 'r', encoding=DEFAULT_ENCODING) as json_file:
        return Session.from_dict(json.load(json_file))


def session_data(data_dir: str) -> Dict[str, Any]:
    """Transform session data to map alphabet letters to likelihood values.

    Args:
        data_dir (str): Directory containing the session data and parameters.

    Returns:
        Dict[str, Any]: Dictionary containing transformed session data with:
            - Mapped alphabet letters to likelihood values
            - Target text from parameters
            - Removed debugging attributes

    Raises:
        FileNotFoundError: If required files are not found in data_dir.
    """
    parameters = load_json_parameters(os.path.join(data_dir,
                                                   DEFAULT_PARAMETERS_FILENAME),
                                      value_cast=True)
    session = read_session(os.path.join(data_dir, SESSION_DATA_FILENAME))
    data = session.as_dict(evidence_only=True)
    data['target_text'] = parameters['task_text']
    return data


@dataclass(frozen=True)
class EvidenceRecord:
    """Record summarizing Inquiry evidence.

    Attributes:
        series (int): Series number.
        inquiry (int): Inquiry number within the series.
        stim (str): Stimulus (letter or icon).
        lm (float): Language model probability.
        eeg (float): EEG evidence value.
        eye (float): Eye tracking evidence value.
        btn (float): Button press evidence value.
        cumulative (float): Cumulative likelihood value.
        inq_position (Optional[int]): Position in inquiry sequence.
        is_target (int): Whether this is the target (1) or not (0).
        presented (int): Whether this was presented (1) or not (0).
        above_threshold (int): Whether above decision threshold (1) or not (0).
    """
    series: int
    inquiry: int
    stim: str
    lm: float
    eeg: float
    eye: float
    btn: float
    cumulative: float
    inq_position: Optional[int]
    is_target: int
    presented: int
    above_threshold: int

    def __iter__(self) -> Iterator[Any]:
        """Iterate over the record's field values.

        Returns:
            Iterator[Any]: Iterator over the record's field values.
        """
        return iter([getattr(self, field.name) for field in fields(self)])


def sqlite_ddl(cls: Any, table_name: str) -> str:
    """Generate SQLite CREATE TABLE statement for a dataclass.

    Args:
        cls (Any): Dataclass to generate DDL for.
        table_name (str): Name of the table to create.

    Returns:
        str: SQLite CREATE TABLE statement.
    """
    conversions = {int: 'integer', str: 'text', float: 'real'}

    column_defs = [
        f'{field.name} {conversions.get(field.type, field.type)}'
        for field in fields(cls)
    ]
    return f"CREATE TABLE {table_name} ({', '.join(column_defs)})"


def sqlite_insert(cls: Any, table_name: str) -> str:
    """Generate SQLite INSERT statement for a dataclass.

    Args:
        cls (Any): Dataclass to generate INSERT for.
        table_name (str): Name of the table to insert into.

    Returns:
        str: SQLite INSERT statement with placeholders.
    """
    placeholders = ['?' for _ in fields(cls)]
    return f"INSERT INTO {table_name} VALUES ({','.join(placeholders)})"


def evidence_records(session: Session) -> List[EvidenceRecord]:
    """Generate evidence records from session data.

    Args:
        session (Session): Session data to process.

    Returns:
        List[EvidenceRecord]: List of evidence records.

    Raises:
        AssertionError: If session has no evidence or no symbol set.
    """
    assert session.has_evidence(
    ), "There is no evidence in the provided session"
    assert session.symbol_set, "Session must define a symbol_set"

    records = []
    for series_index, inquiry_list in enumerate(session.series):
        for inq_index, inquiry in enumerate(inquiry_list):
            evidence = inquiry.stim_evidence(symbol_set=session.symbol_set)
            stimuli = inquiry.stimuli
            # Flatten if stimuli is structured as a nested list.
            if len(stimuli) == 1 and isinstance(stimuli[0], list):
                stimuli = stimuli[0]
            for stim, _likelihood_ev in evidence['likelihood'].items():
                records.append(
                    EvidenceRecord(
                        series=series_index + 1,
                        inquiry=inq_index,
                        stim=stim,
                        lm=evidence['lm_evidence'].get(stim, ''),
                        eeg=evidence['eeg_evidence'].get(stim, ''),
                        eye=evidence.get('eye_evidence', {}).get(stim, ''),
                        btn=evidence.get('btn_evidence', {}).get(stim, ''),
                        cumulative=evidence['likelihood'][stim],
                        inq_position=stimuli.index(stim) if stim in stimuli else None,
                        is_target=int(inquiry.target_letter == stim),
                        presented=int(stim in stimuli),
                        above_threshold=int(evidence['likelihood'][stim] >
                                            session.decision_threshold)))
    return records


def session_db(session: Session, db_file: str = 'session.db') -> None:
    """Create a SQLite database from session data.

    Args:
        session (Session): Session data to store in database.
        db_file (str, optional): Path to database file. Defaults to 'session.db'.

    Database Schema:
        evidence:
        - trial integer (0-based)
        - inquiry integer (0-based)
        - letter text (letter or icon)
        - lm real (language model probability)
        - eeg real (likelihood for the inquiry)
        - btn real (button press evidence)
        - eye real (eyetracker evidence)
        - cumulative real (cumulative likelihood)
        - inq_position integer (inquiry position)
        - is_target integer (boolean)
        - presented integer (boolean)
        - above_threshold integer (boolean)
    """
    # Create database
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    table_name = 'evidence'
    cursor.execute(sqlite_ddl(EvidenceRecord, table_name))
    conn.commit()

    conn.executemany(sqlite_insert(EvidenceRecord, table_name),
                     map(tuple, evidence_records(session)))
    conn.commit()


def session_csv(session: Session, csv_file: str = 'session.csv') -> None:
    """Create a CSV file summarizing session evidence data.

    Args:
        session (Session): Session data to summarize.
        csv_file (str, optional): Path to CSV file. Defaults to 'session.csv'.
    """
    with open(csv_file, "w", encoding=DEFAULT_ENCODING, newline='') as output:
        csv_writer = csv.writer(output, delimiter=',')

        columns = [field.name for field in fields(EvidenceRecord)]
        csv_writer.writerow(columns)
        for record in evidence_records(session):
            csv_writer.writerow(record)


def write_row(excel_sheet: Worksheet, 
              rownum: int, 
              data: Union[EvidenceRecord, List[Any]], 
              background: Optional[PatternFill] = None, 
              border: Optional[Border] = None) -> None:
    """Write a row to an Excel spreadsheet.

    Args:
        excel_sheet (Worksheet): Worksheet to write to.
        rownum (int): Row number to write to.
        data (Union[EvidenceRecord, List[Any]]): Data to write.
        background (Optional[PatternFill], optional): Background fill. Defaults to None.
        border (Optional[Border], optional): Cell border. Defaults to None.
    """
    for col, val in enumerate(data, start=1):
        cell = excel_sheet.cell(row=rownum, column=col)
        cell.value = val
        if background:
            cell.fill = background
        if border:
            cell.border = border


def session_excel(session: Session,
                  excel_file: str = SESSION_SUMMARY_FILENAME,
                  include_charts: bool = True) -> None:
    """Create an Excel spreadsheet summarizing session evidence data.

    Args:
        session (Session): Session data to summarize.
        excel_file (str, optional): Path to Excel file. Defaults to SESSION_SUMMARY_FILENAME.
        include_charts (bool, optional): Whether to include charts. Defaults to True.
    """
    # Define styles and borders to use within the spreadsheet.
    gray_background = PatternFill(start_color='ededed', fill_type='solid')
    white_background = PatternFill(start_color=WHITE, fill_type=None)
    highlighted_background = PatternFill(start_color=YELLOW, fill_type='solid')

    backgrounds_iter = itertools.cycle([gray_background, white_background])

    thin_gray = Side(border_style=BORDER_THIN, color='d3d3d3')
    default_border = Border(left=thin_gray,
                            top=thin_gray,
                            right=thin_gray,
                            bottom=thin_gray)

    thick_black = Side(border_style=BORDER_THIN, color=BLACK)
    new_series_border = Border(left=thin_gray,
                               top=thick_black,
                               right=thin_gray,
                               bottom=thin_gray)

    # Create the workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'session'

    columns = [field.name for field in fields(EvidenceRecord)]
    series, inquiry = None, None

    # Write header
    write_row(sheet, 1, columns)

    # Maps the chart title to the starting row for the data.
    chart_data = {}

    # Write rows
    inq_background = next(backgrounds_iter)
    alp_len = len(session.symbol_set)
    for i, record in enumerate(evidence_records(session)):
        rownum = i + 2  # Excel is 1-indexed; also account for column row.
        border = default_border

        if record.series != series:
            # Place a thick top border before each new series to make it
            # easier to visually scan the spreadsheet.
            series = record.series
            border = new_series_border

        if record.inquiry != inquiry:
            inquiry = record.inquiry
            chart_data[f"Series {series} inquiry {inquiry}"] = rownum

            # Toggle the background for each inquiry for easier viewing.
            inq_background = next(backgrounds_iter)

        # write to spreadsheet
        write_row(sheet,
                  rownum,
                  record,
                  background=highlighted_background
                  if bool(record.is_target) else inq_background,
                  border=border)

    # Add chart for each inquiry
    if include_charts:
        stim_col = columns.index('stim') + 1
        lm_col = columns.index('lm') + 1
        likelihood_col = columns.index('cumulative') + 1

        for title, min_row in chart_data.items():
            max_row = min_row + (alp_len - 1)
            chart = BarChart()
            chart.type = "col"
            chart.title = title
            chart.y_axis.title = 'likelihood'
            chart.x_axis.title = 'stimulus'

            data = Reference(sheet,
                             min_col=lm_col,
                             min_row=min_row,
                             max_row=max_row,
                             max_col=likelihood_col)
            categories = Reference(sheet,
                                   min_col=stim_col,
                                   min_row=min_row,
                                   max_row=max_row)
            chart.add_data(data, titles_from_data=False)
            chart.series[0].title = openpyxl.chart.series.SeriesLabel(v="lm")
            chart.series[1].title = openpyxl.chart.series.SeriesLabel(v="eeg")
            chart.series[2].title = openpyxl.chart.series.SeriesLabel(v="eye")
            chart.series[3].title = openpyxl.chart.series.SeriesLabel(v="btn")
            chart.series[4].title = openpyxl.chart.series.SeriesLabel(
                v="combined")

            chart.set_categories(categories)
            sheet.add_chart(chart, f'M{min_row + 1}')

    # Freeze header row
    sheet.freeze_panes = 'A2'
    workbook.save(excel_file)
    logger.info("Wrote output to %s", excel_file)


if __name__ == "__main__":
    import doctest
    doctest.testmod()
