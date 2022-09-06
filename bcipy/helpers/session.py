"""Helper functions for managing and parsing session.json data."""
import csv
import itertools
import json
import os
import sqlite3
import subprocess
from dataclasses import dataclass, fields
from typing import Any, Dict, List

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import BORDER_THIN, Border, Side
from openpyxl.styles.colors import BLACK, WHITE, YELLOW

from bcipy.config import (
    BCIPY_ROOT,
    DEFAULT_ENCODING,
    DEFAULT_PARAMETER_FILENAME,
    EXPERIMENT_DATA_FILENAME,
    SESSION_DATA_FILENAME,
    SESSION_SUMMARY_FILENAME,
)
from bcipy.helpers.load import (load_experiment_fields, load_experiments,
                                load_json_parameters)
from bcipy.helpers.validate import validate_field_data_written
from bcipy.task.data import Session


def read_session(file_name=SESSION_DATA_FILENAME) -> Session:
    """Read the session data from the given file."""
    with open(file_name, 'r', encoding=DEFAULT_ENCODING) as json_file:
        return Session.from_dict(json.load(json_file))


def session_data(data_dir: str) -> Dict:
    """Returns a dict of session data transformed to map the alphabet letter
    to the likelihood when presenting the evidence. Also removes attributes
    not useful for debugging."""

    parameters = load_json_parameters(os.path.join(data_dir,
                                                   DEFAULT_PARAMETER_FILENAME),
                                      value_cast=True)
    session = read_session(os.path.join(data_dir, SESSION_DATA_FILENAME))
    data = session.as_dict(evidence_only=True)
    data['target_text'] = parameters['task_text']
    return data


def collect_experiment_field_data(experiment_name,
                                  save_path,
                                  file_name=EXPERIMENT_DATA_FILENAME) -> None:
    experiment = load_experiments()[experiment_name]
    experiment_fields = load_experiment_fields(experiment)

    if experiment_fields:
        cmd = (f'python {BCIPY_ROOT}/gui/experiments/ExperimentField.py -e'
               f' "{experiment_name}" -p "{save_path}" -f "{file_name}"')
        subprocess.check_call(cmd, shell=True)
        # verify data was written before moving on
        if not validate_field_data_written(save_path, file_name):
            raise Exception('Field data not collected!')


@dataclass(frozen=True)
class EvidenceRecord:
    """Record summarizing Inquiry evidence."""
    series: int
    inquiry: int
    stim: str
    lm: float
    eeg: float
    btn: float
    cumulative: float
    inq_position: int
    is_target: int
    presented: int
    above_threshold: int

    def __iter__(self):
        return iter([getattr(self, field.name) for field in fields(self)])


def sqlite_ddl(cls: Any, table_name: str) -> str:
    """Sqlite create table statement for the given dataclass"""
    conversions = {int: 'integer', str: 'text', float: 'real'}

    column_defs = [
        f'{field.name} {conversions.get(field.type, field.type)}'
        for field in fields(cls)
    ]
    return f"CREATE TABLE {table_name} ({', '.join(column_defs)})"


def sqlite_insert(cls: Any, table_name: str) -> str:
    """sqlite INSERT statement for the given dataclass."""
    placeholders = ['?' for _ in fields(cls)]
    return f"INSERT INTO {table_name} VALUES ({','.join(placeholders)})"


def evidence_records(session: Session) -> List[EvidenceRecord]:
    """Summarize the session evidence data."""
    assert session.has_evidence(
    ), "There is no evidence in the provided session"
    assert session.symbol_set, "Session must define a symbol_set"

    records = []
    for series_index, inquiry_list in enumerate(session.series):
        for inq_index, inquiry in enumerate(inquiry_list):
            evidence = inquiry.stim_evidence(symbol_set=session.symbol_set)
            stimuli = inquiry.stimuli
            # Flatten if stimuli is structured as a nested list.
            if len(stimuli) == 1 and isinstance(stimuli[0], list):
                stimuli = stimuli[0]
            for stim, _likelihood_ev in evidence['likelihood'].items():
                records.append(
                    EvidenceRecord(
                        series=series_index + 1,
                        inquiry=inq_index,
                        stim=stim,
                        lm=evidence['lm_evidence'].get(stim, ''),
                        eeg=evidence['eeg_evidence'].get(stim, ''),
                        btn=evidence.get('btn_evidence', {}).get(stim, ''),
                        cumulative=evidence['likelihood'][stim],
                        inq_position=stimuli.index(stim) if stim in stimuli else None,
                        is_target=int(inquiry.target_letter == stim),
                        presented=int(stim in stimuli),
                        above_threshold=int(evidence['likelihood'][stim] >
                                            session.decision_threshold)))
    return records


def session_db(session: Session, db_file: str = 'session.db'):
    """Creates a sqlite database from the given session data.

    Parameters
    ----------
        session - task data (evidence values, stim times, etc.)
        db_file - path  of database to write; defaults to session.db

    Database Schema
    ---------------
    evidence:
    - trial integer (0-based)
    - inquiry integer (0-based)
    - letter text (letter or icon)
    - lm real (language model probability for the trial; same for every
        inquiry and only considered in the cumulative value during the
        first inquiry)
    - eeg real (likelihood for the given inquiry; a value of 1.0 indicates
        that the letter was not presented)
    - cumulative real (cumulative likelihood for the trial thus far)
    - inq_position integer (inquiry position; null if not presented)
    - is_target integer (boolean; true(1) if this letter is the target)
    - presented integer (boolean; true if the letter was presented in
        this inquiry)
    - above_threshold (boolean; true if cumulative likelihood was above
        the configured threshold)
    """
    # Create database
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    table_name = 'evidence'
    cursor.execute(sqlite_ddl(EvidenceRecord, table_name))
    conn.commit()

    conn.executemany(sqlite_insert(EvidenceRecord, table_name),
                     map(tuple, evidence_records(session)))
    conn.commit()


def session_csv(session: Session, csv_file='session.csv'):
    """Create a csv file summarizing the evidence data for the given session."""
    with open(csv_file, "w", encoding=DEFAULT_ENCODING, newline='') as output:
        csv_writer = csv.writer(output, delimiter=',')

        columns = [field.name for field in fields(EvidenceRecord)]
        csv_writer.writerow(columns)
        for record in evidence_records(session):
            csv_writer.writerow(record)


def write_row(excel_sheet, rownum, data, background=None, border=None):
    """Helper method to write a row to an Excel spreadsheet"""
    for col, val in enumerate(data, start=1):
        cell = excel_sheet.cell(row=rownum, column=col)
        cell.value = val
        if background:
            cell.fill = background
        if border:
            cell.border = border


def session_excel(session: Session,
                  excel_file=SESSION_SUMMARY_FILENAME,
                  include_charts=True):
    """Create an Excel spreadsheet summarizing the evidence data for the given session."""

    # Define styles and borders to use within the spreadsheet.
    gray_background = PatternFill(start_color='ededed', fill_type='solid')
    white_background = PatternFill(start_color=WHITE, fill_type=None)
    highlighted_background = PatternFill(start_color=YELLOW, fill_type='solid')

    backgrounds_iter = itertools.cycle([gray_background, white_background])

    thin_gray = Side(border_style=BORDER_THIN, color='d3d3d3')
    default_border = Border(left=thin_gray,
                            top=thin_gray,
                            right=thin_gray,
                            bottom=thin_gray)

    thick_black = Side(border_style=BORDER_THIN, color=BLACK)
    new_series_border = Border(left=thin_gray,
                               top=thick_black,
                               right=thin_gray,
                               bottom=thin_gray)

    # Create the workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'session'

    columns = [field.name for field in fields(EvidenceRecord)]
    series, inquiry = None, None

    # Write header
    write_row(sheet, 1, columns)

    # Maps the chart title to the starting row for the data.
    chart_data = {}

    # Write rows
    inq_background = next(backgrounds_iter)
    alp_len = None
    for i, record in enumerate(evidence_records(session)):
        rownum = i + 2  # Excel is 1-indexed; also account for column row.
        border = default_border

        if record.series != series:
            # Place a thick top border before each new series to make it
            # easier to visually scan the spreadsheet.
            series = record.series
            border = new_series_border

        if record.inquiry != inquiry:
            inquiry = record.inquiry

            # Set the alphabet length used for chart data ranges.
            if not alp_len and i > 0:
                alp_len = i
            chart_data[f"Series {series} inquiry {inquiry}"] = rownum

            # Toggle the background for each inquiry for easier viewing.
            inq_background = next(backgrounds_iter)

        # write to spreadsheet
        write_row(sheet,
                  rownum,
                  record,
                  background=highlighted_background
                  if bool(record.is_target) else inq_background,
                  border=border)

    # Add chart for each inquiry
    if include_charts:
        stim_col = columns.index('stim') + 1
        lm_col = columns.index('lm') + 1
        likelihood_col = columns.index('cumulative') + 1

        for title, min_row in chart_data.items():
            max_row = min_row + (alp_len - 1)
            chart = BarChart()
            chart.type = "col"
            chart.title = title
            chart.y_axis.title = 'likelihood'
            chart.x_axis.title = 'stimulus'

            data = Reference(sheet,
                             min_col=lm_col,
                             min_row=min_row,
                             max_row=max_row,
                             max_col=likelihood_col)
            categories = Reference(sheet,
                                   min_col=stim_col,
                                   min_row=min_row,
                                   max_row=max_row)
            chart.add_data(data, titles_from_data=False)
            chart.series[0].title = openpyxl.chart.series.SeriesLabel(v="lm")
            chart.series[1].title = openpyxl.chart.series.SeriesLabel(v="eeg")
            chart.series[2].title = openpyxl.chart.series.SeriesLabel(v="btn")
            chart.series[3].title = openpyxl.chart.series.SeriesLabel(
                v="combined")

            chart.set_categories(categories)
            sheet.add_chart(chart, f'M{min_row + 1}')

    # Freeze header row
    sheet.freeze_panes = 'A2'
    workbook.save(excel_file)
    print("Wrote output to " + excel_file)


if __name__ == "__main__":
    import doctest
    doctest.testmod()
